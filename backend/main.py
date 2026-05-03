from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import os
from datetime import datetime, date
import json
from pathlib import Path
import tempfile
import shutil
from typing import List, Optional
import sys
from dotenv import load_dotenv
import time
from urllib.parse import quote

# === إضافة Supabase Cloud Bridge ===
from supabase import create_client, Client

# Load environment variables from .env file
load_dotenv()

# === إعدادات Supabase ===
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://zfjstzxfeajqlbvsfznm.supabase.co/")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
try:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    print(f"[OK] Supabase connection initialized")
except Exception as e:
    print(f"[WARNING] Supabase connection failed: {e}")
    supabase = None

# Add backend directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from database import SessionLocal, Mission, ImageAnalysis, User, get_db, Base, engine
from ai import process_image, process_zip_file
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import io

app = FastAPI(title="GreenVision API", description="Invasive Plant Monitoring System")

# Initialize database tables on startup
Base.metadata.create_all(bind=engine)
print("[OK] Database tables initialized")

# CORS configuration
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
os.makedirs("uploads", exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

# Load environment variables
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "").strip()

if not OPENROUTER_API_KEY or OPENROUTER_API_KEY == "your-api-key-here":
    print("[WARNING] No valid OpenRouter API key found in .env")
    print("   Warning: No valid OpenRouter API key found in .env")
    print("   Please add your API key: OPENROUTER_API_KEY=sk-or-v1-...")
else:
    print(f"[OK] OpenRouter API key loaded: {OPENROUTER_API_KEY[:20]}...")

AI_MODEL = os.getenv("AI_MODEL", "google/gemini-2.5-flash-lite")

# ==========================================
# دالة مساعدة لرفع الصور للسحابة (The Cloud Bridge)
# ==========================================
import time

def sync_file_to_cloud(local_file_path: str, mission_id: int, plant_data: dict = None):
    """رفع ملف محلي للسحابة مع معلومات النبات المكتشف"""
    if not supabase:
        return None
    
    try:
        file_name = os.path.basename(local_file_path)
        
        # تخطي الملفات غير الصور
        if not file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.gif', '.tiff')):
            return None

        print(f"☁️ جاري مزامنة الملف للسحابة: {file_name}")
        
        # 1. قراءة الملف
        with open(local_file_path, "rb") as f:
            file_bytes = f.read()

        # 2. رفع للمخزن (Storage) مع اسم فريد
        cloud_name = f"mission_{mission_id}_{int(time.time())}_{file_name}"
        supabase.storage.from_("drone-images").upload(
            path=cloud_name,
            file=file_bytes,
            file_options={"content-type": "image/jpeg", "upsert": "true"}
        )
        
        # 3. الحصول على الرابط العام
        public_url = f"{SUPABASE_URL}storage/v1/object/public/drone-images/{cloud_name}"

        # 4. تسجيل في جدول البيانات مع تفاصيل النبات
        detection_record = {
            "image_url": public_url,
            "plant_type": "Pending Analysis...",
            "is_invasive": False,
            "confidence": 0,
            "latitude": 24.7136,
            "longitude": 46.6753
        }
        
        # إذا كانت هناك بيانات نبات، استخدميها
        if plant_data:
            detection_record.update({
                "plant_type": plant_data.get("plant_scientific", "Pending Analysis..."),
                "is_invasive": plant_data.get("is_invasive", False),
                "confidence": plant_data.get("confidence", 0),
                "latitude": plant_data.get("latitude", 24.7136),
                "longitude": plant_data.get("longitude", 46.6753)
            })
        
        supabase.table("detections").insert(detection_record).execute()
        print(f"[OK] Cloud sync successful: {file_name}")
        return public_url

    except Exception as e:
        print(f"[WARNING] Cloud upload failed (local file intact): {e}")
        return None

@app.post("/api/upload")
async def upload_mission(
    background_tasks: BackgroundTasks,
    mission_name: str = Form(...),
    mission_date: str = Form(...),
    notes: str = Form(None),
    files: List[UploadFile] = File(None),
    zip_file: UploadFile = File(None)
):
    """Handle mission upload - Save files FIRST then process in background"""
    
    # DEBUG: Log all received parameters
    print("\n" + "="*60)
    print("DEBUG: /api/upload ENDPOINT RECEIVED:")
    print("="*60)
    print(f"mission_name: '{mission_name}' (type: {type(mission_name).__name__})")
    print(f"mission_date: '{mission_date}' (type: {type(mission_date).__name__})")
    print(f"notes: '{notes}' (type: {type(notes).__name__}, is None: {notes is None})")
    print(f"files: {files} (count: {len(files) if files else 0})")
    print(f"zip_file: {zip_file}")
    print("="*60 + "\n")
    
    db = SessionLocal()
    
    try:
        # 1. Create User/Mission records
        user = db.query(User).filter(User.username == "drone_operator").first()
        if not user:
            user = User(username="drone_operator", email="drone@greenvision.local", role="drone_operator")
            db.add(user)
            db.commit()
            db.refresh(user)
        
        mission = Mission(
            name=mission_name,
            date=mission_date,
            location=notes,
            operator_id=user.id,
            status="processing",
            upload_status="success",
            total_images=0,
            processed_images=0
        )
        db.add(mission)
        db.commit()
        db.refresh(mission)
        
        # DEBUG: Log the notes value
        print(f"✓ Mission created - ID: {mission.id}")
        print(f"  Name: {mission_name}")
        print(f"  Date: {mission_date}")
        print(f"  Notes received: {notes}")
        print(f"  Notes stored in mission.location: {mission.location}")
        
        # 2. SAVE FILES IMMEDIATELY (Synchronously) to PERMANENT DIRECTORY
        # Create permanent directory: uploads/mission_{mission.id}
        mission_dir = os.path.join("uploads", f"mission_{mission.id}")
        os.makedirs(mission_dir, exist_ok=True)
        saved_file_paths = []
        
        # Save individual files
        if files:
            for file in files:
                file_path = os.path.join(mission_dir, file.filename)
                with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                saved_file_paths.append(file_path)
        
        # Save ZIP file
        zip_path = None
        if zip_file and zip_file.filename.endswith('.zip'):
            zip_path = os.path.join(mission_dir, "uploaded.zip")
            with open(zip_path, "wb") as buffer:
                shutil.copyfileobj(zip_file.file, buffer)
        
        # 3. Start Background Task with FILE PATHS (not file objects)
        background_tasks.add_task(
            process_upload_background,
            mission.id,
            mission_dir,         # Pass the mission directory (permanent)
            saved_file_paths,    # Pass list of paths
            zip_path             # Pass zip path
        )
        
        response_data = {
            "success": True,
            "message": "Upload successful. Analysis started.",
            "mission_id": mission.id
        }
        
        print(f"✓ Upload endpoint returning response: {response_data}")
        return response_data
        
    except Exception as e:
        db.rollback()
        print(f"❌ ERROR in upload endpoint: {e}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        
        # If mission was created, mark upload as failed
        if 'mission' in locals():
            try:
                mission.upload_status = "failed"
                db.add(mission)
                db.commit()
            except:
                pass
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")
    finally:
        db.close()

async def process_upload_background(mission_id: int, mission_dir: str, file_paths: list, zip_path: str):
    """Background task to process images from saved paths - ROBUST VERSION"""
    db = SessionLocal()
    
    try:
        mission = db.query(Mission).filter(Mission.id == mission_id).first()
        if not mission:
            return
        
        all_images = file_paths.copy()
        
        # Handle ZIP extraction if exists
        if zip_path and os.path.exists(zip_path):
            extract_dir = os.path.join(mission_dir, "extracted")
            import zipfile
            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(extract_dir)
                
                image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.tiff']
                for root, dirs, files_in_dir in os.walk(extract_dir):
                    for file in files_in_dir:
                        if any(file.lower().endswith(ext) for ext in image_extensions):
                            all_images.append(os.path.join(root, file))
            except Exception as e:
                print(f"✗ Error extracting ZIP: {e}")
        
        # Update total count
        mission.total_images = len(all_images)
        db.commit()
        
        # Process each image
        processed_count = 0
        for image_path in all_images:
            # === الحماية هنا: تعريف المتغيرات بقيم افتراضية قبل الـ Try ===
            plant_status = "unknown"
            plant_scientific = "Unknown"
            plant_common_ar = "غير معروف"
            plant_count = 0
            confidence_val = 0
            is_plant_flag = 0
            gps_data = {}
            
            try:
                # 1. Rate Limiting Protection (Sleep 2s to avoid 429)
                time.sleep(2)

                # 2. Process image with AI
                result = process_image(image_path, OPENROUTER_API_KEY)
                
                # 3. Extract Data safely
                analysis_data = result.get("analysis", {})
                gps_data = result.get("gps", {})
                plants_detected = analysis_data.get("analysis", {}).get("plants_detected", [])
                is_plant_flag = 1 if analysis_data.get("is_plant", False) else 0
                
                if plants_detected:
                    primary_plant = plants_detected[0]
                    plant_scientific = primary_plant.get("scientific_name", "Unknown")
                    enriched = primary_plant.get("enriched_data", {})
                    # Fix: Handle NoneType for Arabic name
                    plant_common_ar = enriched.get("common_name_ar") or primary_plant.get("common_name_ar") or "غير معروف"
                    plant_count = primary_plant.get("count", 1)
                    confidence_val = primary_plant.get("confidence", 0)
                    
                    # Classification Logic - FULL CLASSIFICATION
                    type_from_db = enriched.get("type", "unknown")
                    if type_from_db == "invasive": 
                        plant_status = "invasive"  # غازي
                    elif type_from_db == "native": 
                        plant_status = "native"    # محلي
                    elif type_from_db == "exotic":
                        plant_status = "exotic"    # دخيل/ويبة
                    else: 
                        plant_status = "unknown"

                # 4. Save to Database (Now safe because variables exist)
                image_analysis = ImageAnalysis(
                    mission_id=mission.id,
                    filename=os.path.basename(image_path),
                    original_filename=os.path.basename(image_path),
                    is_plant=is_plant_flag,
                    plant_name_scientific=plant_scientific,
                    plant_name_common_ar=plant_common_ar,
                    plant_count=plant_count,
                    plant_status=plant_status,
                    confidence=confidence_val,
                    latitude=gps_data.get("latitude"),
                    longitude=gps_data.get("longitude"),
                    analysis_result=json.dumps(result, ensure_ascii=False)
                )
                
                db.add(image_analysis)
                processed_count += 1
                mission.processed_images = processed_count
                db.commit()
                print(f"✓ Processed: {os.path.basename(image_path)} -> {plant_status}")
                
                # 5. Sync to Cloud
                plant_info = {
                    "plant_scientific": plant_scientific,
                    "is_invasive": plant_status == "invasive",
                    "confidence": confidence_val,
                    "latitude": gps_data.get("latitude") or 24.7136,
                    "longitude": gps_data.get("longitude") or 46.6753
                }
                sync_file_to_cloud(image_path, mission.id, plant_info)
                
            except Exception as e:
                # لو صار خطأ في صورة واحدة، ما يوقف النظام، يطبع الخطأ ويكمل للي بعدها
                print(f"✗ Error processing {os.path.basename(image_path)}: {e}")
                continue
        
        # Keep status as "processing" so specialist can review in their interface
        # Specialist will see reports automatically under "قيد المعالجة" section
        mission.status = "processing"
        db.commit()
        print(f"✓ Mission {mission_id} processed: {processed_count}/{mission.total_images} images - Ready for specialist review")
        
    except Exception as e:
        # لو صار خطأ كارثي في الدالة كاملة، يحول الحالة إلى Failed عشان ما يعلق
        print(f"CRITICAL Background Error: {e}")
        if 'mission' in locals() and mission:
            mission.status = "failed"
            db.commit()
    finally:
        # NOTE: Do NOT delete mission_dir - files are now permanent!
        db.close()

@app.get("/api/missions")
def get_missions():
    """Get all missions for drone operator sidebar"""
    db = SessionLocal()
    try:
        missions = db.query(Mission).order_by(Mission.created_at.desc()).all()
        
        result = []
        for mission in missions:
            # Count invasive plants for this mission
            invasive_count = db.query(ImageAnalysis).filter(
                ImageAnalysis.mission_id == mission.id,
                ImageAnalysis.is_plant == True,
                ImageAnalysis.plant_status == "invasive"
            ).count()
            
            # Use getattr to safely get upload_status with default value
            upload_status = getattr(mission, 'upload_status', 'success')
            
            result.append({
                "id": mission.id,
                "name": mission.name,
                "date": mission.date,
                "status": mission.status,
                "upload_status": upload_status,
                "total_images": mission.total_images,
                "processed_images": mission.processed_images,
                "invasive_count": invasive_count,
                "created_at": mission.created_at.isoformat() if mission.created_at else None
            })
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.put("/api/missions/{mission_id}/status")
def update_mission_status(mission_id: int, status_data: dict):
    """Update mission status"""
    db = SessionLocal()
    try:
        mission = db.query(Mission).filter(Mission.id == mission_id).first()
        if not mission:
            raise HTTPException(status_code=404, detail="Mission not found")
        
        new_status = status_data.get("status")
        if not new_status:
            raise HTTPException(status_code=400, detail="Status field is required")
        
        mission.status = new_status
        mission.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(mission)
        
        return {
            "id": mission.id,
            "name": mission.name,
            "status": mission.status,
            "message": "Mission status updated successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/stats")
def get_statistics():
    """Get statistics for specialist dashboard"""
    db = SessionLocal()
    try:
        # Total reports
        total_reports = db.query(Mission).count()
        
        # Total images analyzed
        total_images = db.query(ImageAnalysis).count()
        
        # Invasive plants count
        invasive_count = db.query(ImageAnalysis).filter(
            ImageAnalysis.is_plant == True,
            ImageAnalysis.plant_status == "invasive"
        ).count()
        
        # Native plants count
        native_count = db.query(ImageAnalysis).filter(
            ImageAnalysis.is_plant == True,
            ImageAnalysis.plant_status == "native"
        ).count()
        
        # Exotic plants count (الدخيلة/الويبة)
        exotic_count = db.query(ImageAnalysis).filter(
            ImageAnalysis.is_plant == True,
            ImageAnalysis.plant_status == "exotic"
        ).count()
        
        # Pending tasks (processing status)
        pending_count = db.query(Mission).filter(
            Mission.status == "processing"
        ).count()
        
        # Species distribution for chart
        invasive_species = db.query(
            ImageAnalysis.plant_name_scientific,
            ImageAnalysis.plant_name_common_ar
        ).filter(
            ImageAnalysis.is_plant == True,
            ImageAnalysis.plant_status == "invasive",
            ImageAnalysis.plant_name_scientific != ""
        ).distinct().all()
        
        species_distribution = []
        for species in invasive_species:
            count = db.query(ImageAnalysis).filter(
                ImageAnalysis.plant_name_scientific == species[0],
                ImageAnalysis.plant_status == "invasive"
            ).count()
            species_distribution.append({
                "scientific_name": species[0],
                "common_name": species[1],
                "count": count
            })
        
        return {
            "total_reports": total_reports,
            "total_images": total_images,
            "invasive_plants": invasive_count,
            "native_plants": native_count,
            "exotic_plants": exotic_count,
            "pending_tasks": pending_count,
            "species_distribution": species_distribution,
            "success_rate": (invasive_count + native_count + exotic_count) / total_images * 100 if total_images > 0 else 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/mission/{mission_id}")
def get_mission_details(mission_id: int):
    """Get detailed information about a specific mission"""
    db = SessionLocal()
    try:
        mission = db.query(Mission).filter(Mission.id == mission_id).first()
        if not mission:
            raise HTTPException(status_code=404, detail="Mission not found")
        
        # Get all image analyses for this mission
        images = db.query(ImageAnalysis).filter(
            ImageAnalysis.mission_id == mission_id
        ).all()
        
        image_details = []
        for img in images:
            analysis_data = json.loads(img.analysis_result) if img.analysis_result else {}
            image_details.append({
                "id": img.id,
                "filename": img.filename,
                "is_plant": img.is_plant,
                "plant_name_common_ar": img.plant_name_common_ar,
                "plant_name_scientific": img.plant_name_scientific,
                "plant_status": img.plant_status,
                "plant_count": img.plant_count,
                "confidence": img.confidence,
                "latitude": img.latitude,
                "longitude": img.longitude,
                "analysis_result": img.analysis_result,
                "analysis": analysis_data
            })
        
        return {
            "mission": {
                "id": mission.id,
                "name": mission.name,
                "date": mission.date,
                "location": mission.location,
                "status": mission.status,
                "total_images": mission.total_images,
                "processed_images": mission.processed_images
            },
            "images": image_details
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/mission/{mission_id}/download_images")
def download_mission_images(mission_id: int, background_tasks: BackgroundTasks):
    """Download all original images for a mission as ZIP file"""
    db = SessionLocal()
    try:
        mission = db.query(Mission).filter(Mission.id == mission_id).first()
        if not mission:
            raise HTTPException(status_code=404, detail="Mission not found")
        
        # Check if mission directory exists with actual files
        mission_dir = os.path.join("uploads", f"mission_{mission_id}")
        has_files = False
        
        if os.path.exists(mission_dir):
            file_list = []
            for root, dirs, files in os.walk(mission_dir):
                file_list.extend(files)
            has_files = len(file_list) > 0
        
        # If no files found in uploads directory, return 404
        if not has_files:
            raise HTTPException(
                status_code=404, 
                detail=f"No files available for mission {mission_id}. This mission was created before the automatic file storage system was implemented."
            )
        
        # Create ZIP archive from mission directory
        zip_temp_path = os.path.join("uploads", f"mission_{mission_id}_download")
        
        try:
            # Create ZIP file (shutil.make_archive adds .zip extension automatically)
            archive_path = shutil.make_archive(
                base_name=zip_temp_path,
                format='zip',
                root_dir=mission_dir
            )
            
            # Create filename for download with UTF-8 and ASCII-safe versions
            base_filename = f"GreenVision_mission_{mission_id}_{mission.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            
            # For RFC 5987 compliance (handles Arabic/Unicode in Content-Disposition)
            # Use a simple ASCII filename as fallback, and UTF-8 encoded version for clients that support it
            ascii_filename = f"GreenVision_mission_{mission_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            
            print(f"✓ Creating ZIP download: {base_filename}")
            
            # Create proper Content-Disposition header with RFC 5987 encoding
            # This supports both old and new browsers for handling non-ASCII filenames
            try:
                # Try to encode filename as ASCII to check if it needs RFC 5987 encoding
                ascii_filename_only = base_filename.encode('ascii')
                disposition = f'attachment; filename="{base_filename}"'
            except UnicodeEncodeError:
                # If filename contains non-ASCII (Arabic), use RFC 5987 encoding
                utf8_filename = quote(base_filename, safe='')
                disposition = f'attachment; filename="{ascii_filename}"; filename*=UTF-8\'\'{utf8_filename}'
            
            # Schedule cleanup of temporary ZIP file after download completes
            background_tasks.add_task(os.remove, archive_path)
            
            # Return the ZIP file as downloadable response
            return FileResponse(
                path=archive_path,
                filename=ascii_filename,
                media_type="application/zip",
                headers={"Content-Disposition": disposition}
            )
        
        except Exception as e:
            print(f"❌ Error creating ZIP archive: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to create download: {str(e)}")
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Error in download_images endpoint: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/export/{mission_id}")
def export_mission_excel(mission_id: int):
    """Export mission data to Excel file"""
    db = SessionLocal()
    
    try:
        mission = db.query(Mission).filter(Mission.id == mission_id).first()
        if not mission:
            raise HTTPException(status_code=404, detail="Mission not found")
        
        # Get all image analyses
        images = db.query(ImageAnalysis).filter(
            ImageAnalysis.mission_id == mission_id
        ).all()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Mission_{mission_id}"
        
        # Set Arabic font
        arabic_font = Font(name="Arial", size=11)
        
        # Write headers
        headers = [
            "اسم الصورة",
            "نبات؟",
            "الاسم العلمي",
            "غازي \\ محلي",
            "الاسم الشائع",
            "العدد",
            "الثقة",
            "خط العرض",
            "خط الطول",
            "رابط خرائط جوجل",
            "الوصف",
            "طرق المكافحة / الدور البيئي"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(name="Arial", size=12, bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        for row_num, img in enumerate(images, 2):
            # Generate Google Maps link
            maps_link = ""
            if img.latitude and img.longitude:
                maps_link = f"https://maps.google.com/?q={img.latitude},{img.longitude}"
            
            # Get analysis data
            analysis_data = json.loads(img.analysis_result) if img.analysis_result else {}
            plants_detected = analysis_data.get("analysis", {}).get("analysis", {}).get("plants_detected", [])
            
            # Get additional information from all detected plants
            additional_info_parts = []
            if plants_detected:
                for plant in plants_detected:
                    enriched = plant.get("enriched_data", {})
                    plant_type = enriched.get("type", "unknown")
                    
                    if plant_type == "invasive":
                        removal = enriched.get("removal_methods", "")
                        if removal:
                            additional_info_parts.append(f"طرق المكافحة: {removal}")
                    elif plant_type == "native":
                        ecological = enriched.get("ecological_role", "")
                        if ecological:
                            additional_info_parts.append(f"الدور البيئي: {ecological}")
                    elif plant_type == "exotic":
                        # For exotic plants, add descriptive info
                        additional_info_parts.append("نبات دخيل/زينة (غير عدواني)")
            
            additional_info = " | ".join(additional_info_parts)
            
            # Write row data
            ws.cell(row=row_num, column=1, value=img.filename).font = arabic_font
            ws.cell(row=row_num, column=2, value="نعم" if img.is_plant else "لا").font = arabic_font
            ws.cell(row=row_num, column=3, value=img.plant_name_scientific).font = arabic_font
            # Classification: غازي (invasive) / محلي (native) / دخيلة/زينة (exotic) / غير ذلك (only for non-plants)
            classification = ""
            if img.is_plant:
                if img.plant_status == "invasive":
                    classification = "غازي"
                elif img.plant_status == "native":
                    classification = "محلي"
                elif img.plant_status == "exotic":
                    classification = "دخيلة/زينة"
                else:
                    classification = "غير معروف"
            else:
                classification = "غير ذلك"
            ws.cell(row=row_num, column=4, value=classification).font = arabic_font
            ws.cell(row=row_num, column=5, value=img.plant_name_common_ar).font = arabic_font
            ws.cell(row=row_num, column=6, value=img.plant_count)
            ws.cell(row=row_num, column=7, value=f"{img.confidence:.2%}" if img.confidence else "")
            ws.cell(row=row_num, column=8, value=img.latitude)
            ws.cell(row=row_num, column=9, value=img.longitude)
            ws.cell(row=row_num, column=10, value=maps_link).font = Font(color="0000FF", underline="single")
            ws.cell(row=row_num, column=11, 
                   value=analysis_data.get("analysis", {}).get("analysis", {}).get("image_description", "")).font = arabic_font
            ws.cell(row=row_num, column=12, value=additional_info).font = arabic_font
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Return as downloadable file
        filename = f"greenvision_mission_{mission_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/api/reports")
def get_all_reports(status: str = None, date_from: str = None, date_to: str = None):
    """Get all reports with filtering options"""
    db = SessionLocal()
    try:
        query = db.query(Mission)
        
        # Apply filters
        if status:
            query = query.filter(Mission.status == status)
        
        if date_from:
            query = query.filter(Mission.date >= date_from)
        
        if date_to:
            query = query.filter(Mission.date <= date_to)
        
        # Order by date
        missions = query.order_by(Mission.date.desc(), Mission.created_at.desc()).all()
        
        result = []
        for mission in missions:
            # Count invasive plants in this mission
            invasive_count = db.query(ImageAnalysis).filter(
                ImageAnalysis.mission_id == mission.id,
                ImageAnalysis.is_plant == True,
                ImageAnalysis.plant_status == "invasive"
            ).count()
            
            result.append({
                "id": mission.id,
                "name": mission.name,
                "date": mission.date,
                "status": mission.status,
                "total_images": mission.total_images,
                "processed_images": mission.processed_images,
                "invasive_count": invasive_count,
                "created_at": mission.created_at.isoformat() if mission.created_at else None
            })
        
        return result
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()

@app.get("/")
def read_root():
    return {"message": "GreenVision API is running", "version": "1.0.0", "year": 2026}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
